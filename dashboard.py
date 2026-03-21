"""
dashboard.py — Jarvis Job Application Dashboard

Run with:
    streamlit run dashboard.py --server.port 8501

Access via VS Code tunnel at:
    http://localhost:8501
"""

import html
import json
import subprocess
import sys
import time
from datetime import datetime
from pathlib import Path

import openpyxl

import ollama
import streamlit as st
import yaml

# ── Config ────────────────────────────────────────────────────────────────────

BASE_DIR  = Path(__file__).parent
APPS_PATH = BASE_DIR / "data" / "applications.json"
LOG_PATH  = BASE_DIR / "logs" / "agent.log"
CONFIG    = BASE_DIR / "config" / "settings.yaml"
RESUME    = BASE_DIR / "profile" / "resume.tex"
CL_STYLE  = BASE_DIR / "profile" / "cover_letter_template.md"
LETTERS   = BASE_DIR / "data" / "cover_letters"

st.set_page_config(
    page_title="Jarvis",
    page_icon="🦞",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── Styles ────────────────────────────────────────────────────────────────────

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Mono:wght@400;500&family=DM+Sans:wght@300;400;500;600&display=swap');

html, body, [class*="css"] {
    font-family: 'DM Sans', sans-serif;
}
.stApp {
    background: #0f0f0f;
    color: #e8e6e0;
}
[data-testid="stSidebar"] {
    background: #161616;
    border-right: 1px solid #2a2a2a;
}
.metric-card {
    background: #1a1a1a;
    border: 1px solid #2a2a2a;
    border-radius: 12px;
    padding: 1.25rem 1.5rem;
    margin-bottom: 0;
}
.metric-label {
    font-size: 11px;
    font-weight: 500;
    letter-spacing: 0.08em;
    text-transform: uppercase;
    color: #666;
    margin-bottom: 6px;
}
.metric-value {
    font-family: 'DM Mono', monospace;
    font-size: 28px;
    font-weight: 500;
    color: #e8e6e0;
    line-height: 1;
}
.metric-value.green { color: #4ade80; }
.metric-value.amber { color: #fbbf24; }
.metric-value.red   { color: #f87171; }
.job-card {
    background: #1a1a1a;
    border: 1px solid #2a2a2a;
    border-radius: 10px;
    padding: 1rem 1.25rem;
    margin-bottom: 10px;
    transition: border-color 0.15s;
}
.job-card:hover { border-color: #444; }
.job-card.selected { border-color: #4ade80; }
.job-title {
    font-size: 15px;
    font-weight: 600;
    color: #e8e6e0;
    margin-bottom: 3px;
}
.job-meta {
    font-size: 12px;
    color: #888;
    margin-bottom: 8px;
}
.score-pill {
    display: inline-block;
    font-family: 'DM Mono', monospace;
    font-size: 11px;
    font-weight: 500;
    padding: 2px 8px;
    border-radius: 20px;
    margin-right: 6px;
}
.score-high  { background: #14532d; color: #4ade80; }
.score-mid   { background: #713f12; color: #fbbf24; }
.score-low   { background: #450a0a; color: #f87171; }
.log-box {
    background: #0a0a0a;
    border: 1px solid #1e1e1e;
    border-radius: 8px;
    padding: 1rem;
    font-family: 'DM Mono', monospace;
    font-size: 11px;
    color: #888;
    height: 340px;
    overflow-y: auto;
    line-height: 1.7;
}
.log-line.info    { color: #9ca3af; }
.log-line.error   { color: #f87171; }
.log-line.warning { color: #fbbf24; }
.log-line.success { color: #4ade80; }
.section-header {
    font-size: 11px;
    font-weight: 600;
    letter-spacing: 0.1em;
    text-transform: uppercase;
    color: #555;
    padding-bottom: 8px;
    border-bottom: 1px solid #1e1e1e;
    margin-bottom: 16px;
}
.cover-letter-box {
    background: #111;
    border: 1px solid #2a2a2a;
    border-radius: 8px;
    padding: 1.25rem;
    font-size: 13px;
    line-height: 1.8;
    color: #ccc;
    white-space: pre-wrap;
    font-family: 'DM Sans', sans-serif;
}
.reasoning-box {
    background: #1a1a1a;
    border-left: 3px solid #374151;
    padding: 8px 12px;
    border-radius: 0 6px 6px 0;
    font-size: 12px;
    color: #888;
    margin-top: 6px;
    font-style: italic;
}
div[data-testid="stButton"] button {
    border-radius: 8px;
    font-family: 'DM Sans', sans-serif;
    font-weight: 500;
    transition: all 0.15s;
}
</style>
""", unsafe_allow_html=True)


# ── Helpers ───────────────────────────────────────────────────────────────────

@st.cache_data(ttl=5)
def load_applications():
    if not APPS_PATH.exists():
        return []
    with open(APPS_PATH) as f:
        return json.load(f)


def save_applications(apps):
    with open(APPS_PATH, "w") as f:
        json.dump(apps, f, indent=2)
    st.cache_data.clear()


EXCEL_PATH = BASE_DIR / "data" / "applications_tracker.xlsx"
EXCEL_HEADERS = ["Date Applied", "Title", "Company", "Location", "Site", "Score", "ATS Score", "Remote", "URL", "Reasoning"]

def log_to_excel(job: dict) -> None:
    if EXCEL_PATH.exists():
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Applications"
        ws.append(EXCEL_HEADERS)
        # Bold header row
        from openpyxl.styles import Font
        for cell in ws[1]:
            cell.font = Font(bold=True)

    ws.append([
        datetime.utcnow().strftime("%Y-%m-%d"),
        job.get("title", ""),
        job.get("company", ""),
        job.get("location", ""),
        job.get("site", "").capitalize(),
        job.get("score"),
        job.get("ats_score"),
        "Yes" if job.get("is_remote") else "No",
        job.get("job_url", ""),
        job.get("reasoning", ""),
    ])
    wb.save(EXCEL_PATH)


def load_config():
    with open(CONFIG) as f:
        return yaml.safe_load(f)


@st.cache_data(ttl=300)
def load_resume_text():
    if not RESUME.exists():
        return ""
    import re
    raw = RESUME.read_text(encoding="utf-8")
    raw = re.sub(r"\\[a-zA-Z]+\*?(\[.*?\])?\{(.*?)\}", r"\2", raw)
    raw = re.sub(r"\\[a-zA-Z]+", " ", raw)
    raw = re.sub(r"[{}]", " ", raw)
    raw = re.sub(r"%.*", "", raw)
    return re.sub(r"\s+", " ", raw).strip()[:3000]


def load_log_lines(n=80):
    if not LOG_PATH.exists():
        return []
    with open(LOG_PATH, encoding="utf-8", errors="replace") as f:
        return f.readlines()[-n:]


def score_color(score):
    if score is None:
        return "score-low", "?"
    if score >= 8:
        return "score-high", f"{score}/10"
    if score >= 6:
        return "score-mid", f"{score}/10"
    return "score-low", f"{score}/10"


def generate_cover_letter(job, cfg):
    resume   = load_resume_text()
    style    = CL_STYLE.read_text(encoding="utf-8") if CL_STYLE.exists() else ""
    model    = cfg["model"]["name"]

    prompt = f"""Write a tailored cover letter.

CANDIDATE RESUME:
{resume}

COVER LETTER STYLE GUIDE:
{style}

JOB:
Title: {job.get('title','')}
Company: {job.get('company','')}
Location: {job.get('location','')}
Description: {job.get('description','')[:2000]}

Instructions:
- Follow the style guide exactly
- Highlight the most relevant project
- Include one metric from the resume
- 250-350 words
- No date or address header
- Output cover letter text only"""

    response = ollama.chat(
        model=model,
        messages=[{"role": "user", "content": prompt}],
        options={"temperature": 0.4},
    )
    letter = response["message"]["content"].strip()

    humanize_prompt = f"""Rewrite this cover letter to sound like a real person wrote it.
Keep all facts, metrics, and structure identical.
Remove AI giveaways: hollow enthusiasm, corporate filler, perfectly balanced rhythm.
Use contractions. Vary sentence length. Short sentences for emphasis.
Output rewritten letter only.

{letter}"""

    response2 = ollama.chat(
        model=model,
        messages=[{"role": "user", "content": humanize_prompt}],
        options={"temperature": 0.6},
    )
    return response2["message"]["content"].strip()


# ── Sidebar ───────────────────────────────────────────────────────────────────

with st.sidebar:
    st.markdown("### 🦞 Jarvis")
    st.markdown("---")

    _apps    = load_applications()
    _applied = sum(1 for a in _apps if a.get("status") == "applied")
    _queued  = sum(1 for a in _apps if a.get("status") == "auto_apply")
    _review  = sum(1 for a in _apps if a.get("status") == "needs_review")

    st.markdown(f'<div class="metric-card" style="margin-bottom:8px"><div class="metric-label">Applied</div><div class="metric-value green">{_applied}</div></div><div class="metric-card" style="margin-bottom:8px"><div class="metric-label">In queue</div><div class="metric-value amber">{_queued}</div></div><div class="metric-card"><div class="metric-label">Needs review</div><div class="metric-value">{_review}</div></div>', unsafe_allow_html=True)

    st.markdown("---")

    if st.button("Run pipeline now", use_container_width=True):
        with st.spinner("Starting pipeline..."):
            subprocess.Popen(
                [sys.executable, str(BASE_DIR / "run_pipeline.py")],
                cwd=str(BASE_DIR),
            )
        st.success("Pipeline started — check logs")

    if st.button("Refresh data", use_container_width=True):
        st.cache_data.clear()
        st.rerun()


# ── Hub tabs ──────────────────────────────────────────────────────────────────

tab_mainframe, tab_app_agent = st.tabs(["🖥️  Mainframe", "💼  Application Agent"])


# ══════════════════════════════════════════════════════════════════════════════
# TAB: Mainframe
# ══════════════════════════════════════════════════════════════════════════════

with tab_mainframe:
    st.markdown('<div class="section-header">System status</div>', unsafe_allow_html=True)

    # Ollama check
    try:
        ollama.list()
        ollama_status, ollama_color = "Online", "#4ade80"
    except Exception:
        ollama_status, ollama_color = "Offline", "#f87171"

    # Last pipeline run
    _scraped_times = [a.get("scraped_at", "") for a in _apps if a.get("scraped_at")]
    _last_run_raw  = max(_scraped_times) if _scraped_times else None
    _last_run_str  = _last_run_raw[:16].replace("T", " ") + " UTC" if _last_run_raw else "Never"

    col1, col2, col3 = st.columns(3)
    with col1:
        st.markdown(f'<div class="metric-card"><div class="metric-label">Ollama</div><div class="metric-value" style="font-size:18px;color:{ollama_color}">{ollama_status}</div></div>', unsafe_allow_html=True)
    with col2:
        st.markdown(f'<div class="metric-card"><div class="metric-label">Model</div><div class="metric-value" style="font-size:16px;color:#e8e6e0">qwen3:8b</div></div>', unsafe_allow_html=True)
    with col3:
        st.markdown(f'<div class="metric-card"><div class="metric-label">Last pipeline run</div><div class="metric-value" style="font-size:14px;color:#888">{_last_run_str}</div></div>', unsafe_allow_html=True)

    st.markdown("<div style='height:24px'></div>", unsafe_allow_html=True)
    st.markdown('<div class="section-header">Agents</div>', unsafe_allow_html=True)

    agent_col1, agent_col2 = st.columns(2)
    with agent_col1:
        st.markdown('<div class="metric-card"><div style="display:flex;align-items:center;gap:10px"><div style="font-size:22px">💼</div><div><div style="font-weight:600;color:#e8e6e0;margin-bottom:2px">Application Agent</div><div style="font-size:11px;color:#4ade80">Active</div></div></div><div style="font-size:12px;color:#666;margin-top:10px">Scrapes, scores, and queues job applications. Pipeline runs daily at 9am.</div></div>', unsafe_allow_html=True)
    with agent_col2:
        st.markdown('<div class="metric-card" style="opacity:0.4"><div style="display:flex;align-items:center;gap:10px"><div style="font-size:22px">➕</div><div><div style="font-weight:600;color:#e8e6e0;margin-bottom:2px">Add Agent</div><div style="font-size:11px;color:#555">Coming soon</div></div></div><div style="font-size:12px;color:#555;margin-top:10px">Connect a new agent to the hub.</div></div>', unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
# TAB: Application Agent
# ══════════════════════════════════════════════════════════════════════════════

with tab_app_agent:
    import re as _re

    sub_queue, sub_all, sub_logs = st.tabs(["Apply Queue", "All Applications", "Pipeline Logs"])

    # ── Sub-tab: Apply Queue ──────────────────────────────────────────────────

    with sub_queue:
        apps  = load_applications()
        queue = [a for a in apps if a.get("status") == "auto_apply" and not a.get("applied_at")]

        scraped_times = [a.get("scraped_at", "") for a in apps if a.get("scraped_at")]
        last_run = max(scraped_times) if scraped_times else ""
        applied_this_run = sum(
            1 for a in apps
            if a.get("status") == "applied"
            and a.get("applied_at", "") >= last_run
        )

        run_label = "since last run" if last_run else "total"
        st.markdown(
            f'<div class="section-header">Apply queue &nbsp;·&nbsp; {applied_this_run} applied {run_label} &nbsp;·&nbsp; {len(queue)} remaining</div>',
            unsafe_allow_html=True,
        )

        if not queue:
            st.info("Queue is empty — run the pipeline to scrape and score new jobs.")
        else:
            cfg = load_config()

            if "apply_idx" not in st.session_state:
                st.session_state["apply_idx"] = 0

            idx     = min(st.session_state["apply_idx"], len(queue) - 1)
            job     = queue[idx]
            job_key = job["id"][:8]
            ats_key = f"ats_result_{job_key}"
            cl_key  = f"cl_generated_{job_key}"
            if ats_key not in st.session_state:
                st.session_state[ats_key] = None
            if cl_key not in st.session_state:
                st.session_state[cl_key] = None

            pill_class, pill_text = score_color(job.get("score"))
            ats_score = job.get("ats_score")
            ats_color = "#4ade80" if (ats_score or 0) >= 70 else "#fbbf24" if (ats_score or 0) >= 50 else "#f87171"

            _title    = html.escape(job.get('title', '—'))
            _company  = html.escape(job.get('company', '—'))
            _location = html.escape(job.get('location', '—'))
            _site     = html.escape(job.get('site', '—').capitalize())
            _remote   = '<div style="margin-top:6px"><span style="font-size:11px;color:#4ade80;font-family:DM Mono,monospace">Remote</span></div>' if job.get('is_remote') else ''
            _ats_div  = f'<div style="font-family:DM Mono,monospace;font-size:12px;color:{ats_color};margin-top:6px">ATS {ats_score}/100</div>' if ats_score is not None else ''
            st.markdown(
                f'<div class="job-card" style="padding:1.5rem;margin-bottom:12px">'
                f'<div style="display:flex;justify-content:space-between;align-items:flex-start">'
                f'<div>'
                f'<div class="job-title" style="font-size:20px;margin-bottom:6px">{_title}</div>'
                f'<div class="job-meta" style="font-size:13px">{_company} &nbsp;·&nbsp; {_location} &nbsp;·&nbsp; {_site}</div>'
                f'{_remote}'
                f'</div>'
                f'<div style="text-align:right;flex-shrink:0;padding-left:16px">'
                f'<span class="score-pill {pill_class}" style="font-size:13px;padding:4px 12px">{pill_text}</span>'
                f'{_ats_div}'
                f'</div>'
                f'</div>'
                f'</div>',
                unsafe_allow_html=True,
            )

            if job.get("reasoning"):
                st.markdown(f'<div class="reasoning-box">{job["reasoning"]}</div>', unsafe_allow_html=True)
            if job.get("ats_score_reasoning"):
                st.markdown(f'<div class="reasoning-box" style="border-left-color:#4ade80;margin-top:6px">ATS: {job["ats_score_reasoning"]}</div>', unsafe_allow_html=True)

            st.markdown("<div style='height:20px'></div>", unsafe_allow_html=True)

            url = job.get("job_url", "")
            c1, c2, c3 = st.columns([2, 1, 1])

            with c1:
                if url:
                    st.link_button("Open & Apply", url, use_container_width=True, type="primary")
                else:
                    st.button("No URL available", use_container_width=True, disabled=True)

            with c2:
                if st.button("Mark applied", key=f"apply_{job_key}", use_container_width=True, type="primary"):
                    for a in apps:
                        if a["id"] == job["id"]:
                            a["status"]     = "applied"
                            a["applied_at"] = datetime.utcnow().isoformat()
                            cl = st.session_state.get(cl_key)
                            if cl:
                                LETTERS.mkdir(exist_ok=True)
                                company_safe = _re.sub(r'[^\w]', '_', job.get('company', 'co'))[:30]
                                title_safe   = _re.sub(r'[^\w]', '_', job.get('title', 'role'))[:30]
                                name = f"{datetime.utcnow().strftime('%Y%m%d')}_{company_safe}_{title_safe}.txt"
                                (LETTERS / name).write_text(cl, encoding="utf-8")
                                a["cover_letter_path"] = str(LETTERS / name)
                            break
                    save_applications(apps)
                    log_to_excel(job)
                    st.session_state["apply_idx"] = idx
                    st.rerun()

            with c3:
                if st.button("Skip", key=f"skip_{job_key}", use_container_width=True):
                    for a in apps:
                        if a["id"] == job["id"]:
                            a["status"] = "skipped"
                            break
                    save_applications(apps)
                    st.session_state["apply_idx"] = idx
                    st.rerun()

            st.markdown("---")

            t1, t2 = st.columns(2)

            with t1:
                if st.button("Full ATS report", key=f"ats_{job_key}", use_container_width=True):
                    with st.spinner("Running full ATS analysis..."):
                        try:
                            sys.path.insert(0, str(BASE_DIR))
                            from agents.ats_scanner import scan_job
                            st.session_state[ats_key] = scan_job(job, quick=False)
                        except Exception as e:
                            st.error(f"ATS scan failed: {e}")

            with t2:
                if st.button("Generate cover letter", key=f"gen_{job_key}", use_container_width=True):
                    with st.spinner("Writing + humanizing..."):
                        try:
                            letter = generate_cover_letter(job, cfg)
                            st.session_state[cl_key] = letter
                        except Exception as e:
                            st.error(f"Generation failed: {e}")

            if st.session_state.get(ats_key):
                ats = st.session_state[ats_key]
                if "error" in ats:
                    st.error(f"ATS scan error: {ats['error']}")
                else:
                    score     = ats.get("ats_score", 0)
                    bar_color = "#4ade80" if score >= 70 else "#fbbf24" if score >= 50 else "#f87171"

                    with st.expander(f"ATS Report — {score}/100", expanded=True):
                        st.markdown(f'<div style="margin-bottom:12px"><div style="font-family:\'DM Mono\',monospace;font-size:22px;font-weight:500;color:{bar_color}">{score}/100</div><div style="background:#1e1e1e;border-radius:4px;height:6px;margin:6px 0;overflow:hidden"><div style="background:{bar_color};width:{score}%;height:100%;border-radius:4px"></div></div><div style="font-size:12px;color:#888;font-style:italic">{ats.get("score_reasoning","")}</div></div>', unsafe_allow_html=True)

                        verdict = ats.get("overall_verdict", "")
                        if verdict:
                            st.markdown(f'<div style="background:#1a1a1a;border-left:3px solid #374151;padding:10px 14px;border-radius:0 8px 8px 0;font-size:13px;color:#aaa;margin-bottom:14px">{verdict}</div>', unsafe_allow_html=True)

                        missing = ats.get("missing_keywords", [])
                        present = ats.get("present_keywords", [])
                        col_m, col_p = st.columns(2)

                        with col_m:
                            st.markdown("**Missing keywords**")
                            if missing:
                                for k in missing:
                                    imp_color = "#f87171" if k.get("importance") == "high" else "#fbbf24" if k.get("importance") == "medium" else "#888"
                                    st.markdown(f'<div style="display:flex;align-items:flex-start;gap:8px;margin-bottom:6px"><span style="font-family:DM Mono,monospace;font-size:11px;background:#1e1e1e;padding:2px 7px;border-radius:4px;color:{imp_color};white-space:nowrap">{k.get("keyword","")}</span><span style="font-size:11px;color:#666">{k.get("where_to_add","")}</span></div>', unsafe_allow_html=True)
                            else:
                                st.markdown('<span style="font-size:12px;color:#4ade80">No critical gaps</span>', unsafe_allow_html=True)

                        with col_p:
                            st.markdown("**Already covered**")
                            if present:
                                chips = " ".join(f'<span style="font-family:DM Mono,monospace;font-size:11px;background:#14532d;color:#4ade80;padding:2px 7px;border-radius:4px;margin:2px;display:inline-block">{k}</span>' for k in present[:10])
                                st.markdown(chips, unsafe_allow_html=True)

                        weak = ats.get("weak_bullets", [])
                        if weak:
                            st.markdown("**Bullet rewrites**")
                            for b in weak:
                                st.markdown(f'<div style="background:#1a1a1a;border-radius:8px;padding:10px 12px;margin-bottom:8px"><div style="font-size:11px;color:#f87171;margin-bottom:4px">Before</div><div style="font-size:12px;color:#888;margin-bottom:8px">{b.get("original","")}</div><div style="font-size:11px;color:#4ade80;margin-bottom:4px">After</div><div style="font-size:12px;color:#ccc">{b.get("rewrite","")}</div></div>', unsafe_allow_html=True)

                        wins = ats.get("quick_wins", [])
                        if wins:
                            st.markdown("**Quick wins (under 10 min)**")
                            for w in wins:
                                st.markdown(f'<div style="font-size:12px;color:#ccc;padding:4px 0;border-bottom:1px solid #1e1e1e">- {w}</div>', unsafe_allow_html=True)

                        issues = ats.get("skills_issues", [])
                        if issues:
                            st.markdown("**Skills section issues**")
                            for iss in issues:
                                st.markdown(f'<div style="font-size:12px;color:#fbbf24;margin-bottom:4px">! {iss.get("issue","")} — <span style="color:#888">{iss.get("fix","")}</span></div>', unsafe_allow_html=True)

            if st.session_state.get(cl_key):
                with st.expander("Cover letter — click to expand/edit", expanded=True):
                    edited = st.text_area(
                        "Edit before submitting",
                        value=st.session_state[cl_key],
                        height=320,
                        key=f"cl_edit_{job_key}",
                        label_visibility="collapsed",
                    )
                    st.session_state[cl_key] = edited

                    col_a, col_b = st.columns(2)
                    with col_a:
                        if st.button("Copy to clipboard", key=f"copy_{job_key}"):
                            st.code(edited, language=None)
                    with col_b:
                        if st.button("Save edits", key=f"save_{job_key}"):
                            LETTERS.mkdir(exist_ok=True)
                            company_safe = _re.sub(r'[^\w]', '_', job.get('company', 'co'))[:30]
                            title_safe   = _re.sub(r'[^\w]', '_', job.get('title', 'role'))[:30]
                            name = f"{datetime.utcnow().strftime('%Y%m%d')}_{company_safe}_{title_safe}.txt"
                            (LETTERS / name).write_text(edited, encoding="utf-8")
                            st.success("Saved.")

    # ── Sub-tab: All Applications ─────────────────────────────────────────────

    with sub_all:
        st.markdown('<div class="section-header">All applications</div>', unsafe_allow_html=True)

        apps = load_applications()

        status_filter = st.selectbox(
            "Filter by status",
            ["All", "applied", "auto_apply", "needs_review", "skipped", "rejected", "scraped"],
        )

        filtered = apps if status_filter == "All" else [a for a in apps if a.get("status") == status_filter]
        filtered = sorted(filtered, key=lambda x: x.get("score") or 0, reverse=True)

        st.markdown(f"**{len(filtered)} jobs**")
        st.markdown("---")

        for job in filtered:
            pill_class, pill_text = score_color(job.get("score"))
            status = job.get("status", "unknown")
            status_colors = {
                "applied":      "#14532d",
                "auto_apply":   "#713f12",
                "needs_review": "#1e3a5f",
                "skipped":      "#1f1f1f",
                "rejected":     "#450a0a",
                "scraped":      "#1a1a1a",
            }
            bg = status_colors.get(status, "#1a1a1a")

            _title     = html.escape(job.get('title', '—'))
            _company   = html.escape(job.get('company', '—'))
            _location  = html.escape(job.get('location', '—'))
            _reasoning = html.escape(job.get('reasoning', '')) if job.get('reasoning') else ''
            _reasoning_html = f'<div class="reasoning-box">{_reasoning}</div>' if _reasoning else ''
            st.markdown(
                f'<div class="job-card" style="border-left:3px solid {bg};padding-left:14px">'
                f'<div style="display:flex;justify-content:space-between;align-items:flex-start">'
                f'<div><div class="job-title">{_title}</div><div class="job-meta">{_company} · {_location}</div></div>'
                f'<div style="text-align:right"><span class="score-pill {pill_class}">{pill_text}</span>'
                f'<div style="font-size:10px;color:#555;margin-top:4px;font-family:DM Mono,monospace">{status}</div></div>'
                f'</div>{_reasoning_html}</div>',
                unsafe_allow_html=True,
            )

    # ── Sub-tab: Pipeline Logs ────────────────────────────────────────────────

    with sub_logs:
        st.markdown('<div class="section-header">Pipeline logs</div>', unsafe_allow_html=True)

        auto_refresh = st.checkbox("Auto-refresh every 5 seconds", value=False)

        lines = load_log_lines(100)
        log_html = ""
        for line in lines:
            line = line.strip()
            if not line:
                continue
            if "[ERROR]" in line:
                css = "error"
            elif "[WARNING]" in line:
                css = "warning"
            elif any(w in line for w in ["complete", "saved", "Applied", "✓"]):
                css = "success"
            else:
                css = "info"
            safe = line.replace("<", "&lt;").replace(">", "&gt;")
            log_html += f'<div class="log-line {css}">{safe}</div>\n'

        st.markdown(f'<div class="log-box">{log_html}</div>', unsafe_allow_html=True)

        if auto_refresh:
            time.sleep(5)
            st.rerun()

        if st.button("Refresh logs"):
            st.rerun()
